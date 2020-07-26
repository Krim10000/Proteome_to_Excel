
import re
#import pandas as pd
import openpyxl
from openpyxl import Workbook
#import sre_yield

wb = Workbook()
ws =  wb.active
ws.title = "Proteinas"
name="Proteinas.xlsx"


#Get new files in
#https://www.uniprot.org/proteomes/

Protein_filename = "UP000000625_83333.fasta"
G=open(Protein_filename, "r")

Check =G.read()

#    faces = [int(face) for face in faces_txt]
#SV=SequenceVersion
start = 'SV=\d' # SV= + numero
#https://www.uniprot.org/help/fasta-headers
end = '>sp|>tr'
#sp
#>db|UniqueIdentifier|EntryName ProteinName OS=OrganismName OX=OrganismIdentifier [GN=GeneName ]PE=ProteinExistence SV=SequenceVersion
###
#tr
#>db|UniqueIdentifier archived from Release ReleaseNumber ReleaseDate SV=SequenceVersion

def posciones(secuencia):

    minicio = re.finditer(secuencia ,Check)
    pinicio = [match.start() for match in minicio]
    return pinicio

Lstart=posciones(start)

Lend=posciones(end)

#Creacion de listas
Lseq=[]
Ldesq=[]
for x in range(len(Lstart)):
    for y in range(len(Lend)):
        if Lend[y]>Lstart[x]:
            Lseq.append(Check[Lstart[x]+4:Lend[y]])
            Ldesq.append(Check[Lend[y-1]:Lstart[x]+4])
            
            break
#ultimo elemento        
Lseq.append(Check[Lstart[-1]+4:len(Check)])
Ldesq.append(Check[Lend[-1]:Lstart[-1]])
#Escritura en excel

ws.cell(row=1, column=1).value="Descripción"
ws.cell(row=1, column=2).value="Secuencia"
ws.cell(row=1, column=3).value="Largo"

for x in range(len(Lseq)):
    #remueve espacios vacios y nuevas lineas.
    Lseq[x] = re.sub(r"[\n\t\s]*", "", Lseq[x])
    ws.cell(row=(x+2), column=1).value=Ldesq[x]
    ws.cell(row=(x+2), column=2).value=Lseq[x]
    ws.cell(row=(x+2), column=3).value=len(Lseq[x])
    
wb.save(filename = name)

